from flask import Blueprint, jsonify, request
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

form_bp = Blueprint('form', __name__)

# File paths
EXCEL_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'submissions.xlsx')
COUNTER_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'counter.json')

def ensure_data_directory():
    """Ensure the data directory exists"""
    data_dir = os.path.dirname(EXCEL_FILE)
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

def initialize_excel_file():
    """Initialize Excel file with headers if it doesn't exist"""
    ensure_data_directory()
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Submissions"
        
        # Set headers
        headers = ['ID', 'Name', 'Phone', 'Email', 'School Name', 'Selected Robot Types', 'Submitted At']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(EXCEL_FILE)

def get_counter():
    """Get current counter from file"""
    ensure_data_directory()
    if os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, 'r') as f:
            data = json.load(f)
            return data.get('count', 0), data.get('max_submissions', 10)
    return 0, 10

def update_counter(count):
    """Update counter in file"""
    ensure_data_directory()
    data = {'count': count, 'max_submissions': 10}
    with open(COUNTER_FILE, 'w') as f:
        json.dump(data, f)

def add_submission_to_excel(name, phone, email, school_name, selected_robots):
    """Add a new submission to the Excel file"""
    initialize_excel_file()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Get next row number
    next_row = ws.max_row + 1
    submission_id = next_row - 1  # Subtract 1 because first row is headers
    
    # Convert robot list to comma-separated string
    robots_str = ', '.join(selected_robots) if isinstance(selected_robots, list) else selected_robots
    
    # Add data
    ws.cell(row=next_row, column=1, value=submission_id)
    ws.cell(row=next_row, column=2, value=name)
    ws.cell(row=next_row, column=3, value=phone)
    ws.cell(row=next_row, column=4, value=email)
    ws.cell(row=next_row, column=5, value=school_name)
    ws.cell(row=next_row, column=6, value=robots_str)
    ws.cell(row=next_row, column=7, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    
    wb.save(EXCEL_FILE)
    return submission_id

def get_all_submissions():
    """Get all submissions from Excel file"""
    if not os.path.exists(EXCEL_FILE):
        return []
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    submissions = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] is not None:  # Check if row has data
            submissions.append({
                'id': row[0],
                'name': row[1],
                'phone': row[2],
                'email': row[3],
                'school_name': row[4],
                'selected_robots': row[5],
                'submitted_at': row[6]
            })
    
    return submissions

@form_bp.route('/submit', methods=['POST'])
def submit_form():
    # Get current counter
    count, max_submissions = get_counter()
    
    # Check if submission limit is reached
    if count >= max_submissions:
        return jsonify({
            'success': False,
            'message': 'Form submission limit reached. No more submissions allowed.',
            'counter': {
                'count': count,
                'max_submissions': max_submissions,
                'submissions_remaining': max_submissions - count,
                'is_limit_reached': True
            }
        }), 403
    
    # Get form data
    data = request.json
    if not data:
        return jsonify({
            'success': False,
            'message': 'No data provided'
        }), 400
    
    # Validate required fields
    required_fields = ['name', 'phone', 'school_name', 'selected_robots']
    for field in required_fields:
        if not data.get(field):
            return jsonify({
                'success': False,
                'message': f'Missing required field: {field}'
            }), 400
    
    # Validate name contains only letters and spaces
    import re
    name_pattern = re.compile(r'^[A-Za-z\s]+$')
    if not name_pattern.match(data['name']):
        return jsonify({
            'success': False,
            'message': 'Name must contain only letters and spaces'
        }), 400
    
    # Validate phone number is exactly 8 digits
    phone_pattern = re.compile(r'^[0-9]{8}$')
    if not phone_pattern.match(data['phone']):
        return jsonify({
            'success': False,
            'message': 'Phone number must be exactly 8 digits'
        }), 400
    
    # Validate email contains @ symbol if provided
    email = data.get('email', '').strip()
    if email and '@' not in email:
        return jsonify({
            'success': False,
            'message': 'Email must contain @ symbol'
        }), 400
    
    # Validate selected_robots is a list and contains valid robot types
    selected_robots = data['selected_robots']
    if not isinstance(selected_robots, list) or len(selected_robots) == 0:
        return jsonify({
            'success': False,
            'message': 'Please select at least one robot type'
        }), 400
    
    allowed_robots = ['Cleaning Robot', 'Security Robot', 'Delivery Robot']
    for robot in selected_robots:
        if robot not in allowed_robots:
            return jsonify({
                'success': False,
                'message': f'Invalid robot type selected: {robot}'
            }), 400
    
    # Add submission to Excel file
    submission_id = add_submission_to_excel(
        data['name'],
        data['phone'],
        data.get('email', ''),  # Optional field
        data['school_name'],
        selected_robots
    )
    
    # Increment counter
    count += 1
    update_counter(count)
    
    return jsonify({
        'success': True,
        'message': 'Form submitted successfully!',
        'submission': {
            'id': submission_id,
            'name': data['name'],
            'phone': data['phone'],
            'email': data.get('email', ''),
            'school_name': data['school_name'],
            'selected_robots': selected_robots,
            'submitted_at': datetime.now().isoformat()
        },
        'counter': {
            'count': count,
            'max_submissions': max_submissions,
            'submissions_remaining': max_submissions - count,
            'is_limit_reached': count >= max_submissions
        }
    }), 201

@form_bp.route('/status', methods=['GET'])
def get_status():
    count, max_submissions = get_counter()
    
    return jsonify({
        'counter': {
            'count': count,
            'max_submissions': max_submissions,
            'submissions_remaining': max_submissions - count,
            'is_limit_reached': count >= max_submissions
        }
    })

@form_bp.route('/submissions', methods=['GET'])
def get_submissions():
    submissions = get_all_submissions()
    count, max_submissions = get_counter()
    
    return jsonify({
        'submissions': submissions,
        'counter': {
            'count': count,
            'max_submissions': max_submissions,
            'submissions_remaining': max_submissions - count,
            'is_limit_reached': count >= max_submissions
        }
    })

@form_bp.route('/reset', methods=['POST'])
def reset_counter():
    try:
        data = request.get_json()
        
        # Check if password is provided
        if not data or 'password' not in data:
            return jsonify({
                'success': False,
                'message': 'Password is required'
            }), 400
        
        # Validate password
        if data['password'] != '25485650':
            return jsonify({
                'success': False,
                'message': 'Incorrect password'
            }), 401
        
        # Reset counter
        update_counter(0)
        
        return jsonify({
            'success': True,
            'message': 'Counter reset successfully!',
            'counter': {
                'count': 0,
                'max_submissions': 10,
                'submissions_remaining': 10,
                'is_limit_reached': False
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Error resetting counter'
        }), 500

@form_bp.route('/download', methods=['GET'])
def download_excel():
    """Download the Excel file with all submissions"""
    if os.path.exists(EXCEL_FILE):
        from flask import send_file
        return send_file(EXCEL_FILE, as_attachment=True, download_name='form_submissions.xlsx')
    else:
        return jsonify({
            'success': False,
            'message': 'No submissions file found'
        }), 404

